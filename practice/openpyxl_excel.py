import openpyxl

# --------------讀取存檔--------------

# 讀取Excel檔案
workbook = openpyxl.load_workbook('OwO.xlsx')

# 存取Excel檔案
workbook.save('OwO.xlsx')

# --------------工作表操作--------------

# 查看全部工作表
print(workbook.sheetnames)    # Output: ['工作表1']

# 選取特定工作表
print(workbook['工作表1'])   # Output: <Worksheet "工作表1">

# 最大列數
mxR = sheet.max_row

# 最大行數
mxC = sheet.max_row

# 新增工作表(放在最後方)
workbook.create_sheet('工作表2')  # Output: <Worksheet "工作表2">

# --------------單個儲存格操作--------------

# 直接選取-方法一
sheet = workbook['工作表1']
print(sheet['B3'].value)   # Output: TwT

# 直接選取-方法二
sheet = workbook['工作表1']
print(sheet.cell(row=3, column=2).value)   # Output: TwT

# 選取後，直接賦值修改值
sheet = workbook['工作表1']
sheet.cell(row=3, column=2).value = 'OwO'
print(sheet.cell(row=3, column=2).value)   # Output: OwO