import openpyxl
import requests
from bs4 import BeautifulSoup

# 將匯率爬蟲包為一個函式，以幣種作為輸入、匯率作為回傳值
def crawler(cointype):
    # 建立請求的URL(=欲爬取網站)
    url = f'https://www.google.com/search?q={cointype}'
    # 設定請求標頭(模擬正常的瀏覽器行為)
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
    }
    # 發送請求並獲取回應
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'html.parser')
    ele = soup.find('span', class_='DFlfde SwHCTb')

    if ele:
        return ele.text
    else:
        return None

# 讀入目標Excel
workbook = openpyxl.load_workbook('匯率即時更新.xlsx')
# 讀取目標分頁
sheet = workbook['即時匯率']
# 讀取目前row的最大值
mxR = sheet.max_row
# print(mxR)      # Output: 4

cointypes = []

# 讀取指定單元格，並將其加入一串列
for r in range(2, mxR + 1):
    cointype = sheet.cell(row=r, column=1).value
    cointypes.append(cointype)

# print(cointypes)   # Output: ['美金', '日幣', '歐元']

# 讀取表單部分完成後，再將匯率爬蟲與此結合
nowrow = 2
for cointype in cointypes:
    exchange_rate = crawler(cointype)
    sheet.cell(row=nowrow, column=2).value = exchange_rate
    nowrow += 1

workbook.save('匯率即時更新.xlsx')

print('已完成更新【匯率即時更新.xlsx】')